import os
import pandas as pd
import glob
import datetime
import csv
import openpyxl
import PySimpleGUI as sg
import matplotlib.pyplot as plt
import pickle
Default_Date_today   = datetime.date.today()
if not os.path.exists('C:\LV_restricted_Folder'):    
    layout = [[sg.Text('Enter SCRIPT Checksum Value:',   size=(40,1)), sg.InputText()],
              [sg.Text('Enter PARAMETER Checksum Value:',size=(40,1)), sg.InputText()],              
              [sg.Text('Enter SPREAD min Station Number:',size=(40,1)), sg.InputText()],
              [sg.Text('Enter SPREAD max Station Number:',size=(40,1)), sg.InputText()],
              [sg.Submit(), sg.Cancel()]
             ]
    window = sg.Window('LV Script/Parameter_Checksum input',
            auto_size_text=True, default_element_size=(10, 1)).Layout(layout)
    event, values = window.Read()
    if event is None or event == 'Cancel':
        sg.PopupAutoClose('Exiting LV Script-Parameter Input',line_width=60)
        os._exit(1)
    window.Close()
    os.makedirs('C:\LV_Restricted_Folder')
    Script_Checksum     = values[0]
    Parameter_Checksum  = values[1]
    SPREAD_min_Station  = int(values[2])
    SPREAD_max_Station  = int(values[3])    
    pickle_dict = {1:Script_Checksum,2:Parameter_Checksum,3:SPREAD_min_Station,4:SPREAD_max_Station}
    pickle_out  = open("C:\LV_Restricted_Folder\LV_fixed_parameters","wb")
    pickle.dump(pickle_dict,pickle_out)
    pickle_out.close()

elif os.path.exists('C:\LV_Restricted_Folder\LV_fixed_parameters'):
    Pickle_in= open("C:\LV_Restricted_Folder\LV_fixed_parameters","rb")
    pickle_dict         = pickle.load(Pickle_in)
    Script_Checksum     = pickle_dict[1]
    Parameter_Checksum  = pickle_dict[2]
    SPREAD_min_Station  = pickle_dict[3]
    SPREAD_max_Station  = pickle_dict[4]

else:
    Script_Checksum     = []
    Parameter_Checksum  = []
    SPREAD_min_Station  = 0
    SPREAD_max_Station  = 20000

layout = [[sg.Text('Enter Desired Deployment Age (Days) (Default = 10 Days) :',  size=(45, 1)), sg.Slider(range=(1, 50), orientation='h', size=(10, 15), default_value=10)],
          [sg.Text('Enter Desired Impedence Tolarance Value (Default = 0) :',   size=(45, 1)), sg.Slider(range=(0, 50), orientation='h', size=(10, 15), default_value=0)],
          [sg.Text('Enter Desired BattVoltage Value (Default = 16.0) :',         size=(45, 1)), sg.InputText(16.0)],
          [sg.Text('Enter High Impedence Value for GSRX3 (Default = 9465) :', size=(45, 1)), sg.InputText(9465)],
          [sg.Text('Enter Low Impedence Value for GSRX3 (Default = 7128) :', size=(45, 1)), sg.InputText(7128)],
          [sg.Text('Enter High Impedence Value for GSR-1C (Default = 750) :', size=(45, 1)), sg.InputText(750)],
          [sg.Text('Enter Low Impedence Value for GSR-1C (Default = 950) :', size=(45, 1)), sg.InputText(950)],
          [sg.Text('Enter Date as YYYY-MM-DD (Default Today) :', size=(45, 1)), sg.InputText(Default_Date_today)],
          [sg.Text('Enter Receiver Station increment (Default = 1) :', size=(45, 1)), sg.InputText(1)],
          [sg.Text('Enter Spread min Station Number :',size=(45,1)), sg.InputText(SPREAD_min_Station)],
          [sg.Text('Enter Spread max Station Number :',size=(45,1)), sg.InputText(SPREAD_max_Station)],
          [sg.Text('Enter Script Checksum Value:', size=(45, 1)), sg.InputText(Script_Checksum)],
          [sg.Text('Enter Parameter Checksum Value:', size=(45, 1)), sg.InputText(Parameter_Checksum)],             
          [sg.Submit(),sg.Cancel()]
         ]

window = sg.Window('Please Input LineViewer QC Parameters Limit:',auto_size_text=True, default_element_size=(10, 1)).Layout(layout)      
event, values = window.Read()

if event is None or event == 'Cancel':
        sg.PopupAutoClose('Exiting LV QC',line_width=60)
        os._exit(1)

SPREAD_min_Station  = int(values[9])
SPREAD_max_Station  = int(values[10])
Script_Checksum     = values[11]
Parameter_Checksum  = values[12]
pickle_dict = {1:Script_Checksum,2:Parameter_Checksum,3:SPREAD_min_Station,4:SPREAD_max_Station}
pickle_out  = open("C:\LV_Restricted_Folder\LV_fixed_parameters","wb")
pickle.dump(pickle_dict,pickle_out)
pickle_out.close()

Deployment_Age                  = float(values[0])
Percent_of_Threshold_Impedence  = float(values[1])
DesiredBattVoltage              = float(values[2])
High_Impedence_GSRX3            = float(values[3])
Low_Impedence_GSRX3             = float(values[4])
High_Impedence_GSR_1C           = float(values[5])
Low_Impedence_GSR_1C            = float(values[6])
Default_Date                    = pd.to_datetime(values[7])
Rec_Station_increment           = int(values[8])

window.Close()

Low_Threshold_Imp_GSRX3  = Low_Impedence_GSRX3 - (Low_Impedence_GSRX3*(Percent_of_Threshold_Impedence)/100)
High_Threshold_Imp_GSRX3 = High_Impedence_GSRX3 + (High_Impedence_GSRX3*(Percent_of_Threshold_Impedence)/100)

Low_Threshold_Imp_GSR_1C  = Low_Impedence_GSR_1C - (Low_Impedence_GSR_1C*(Percent_of_Threshold_Impedence)/100)
High_Threshold_Imp_GSR_1C = High_Impedence_GSR_1C + (High_Impedence_GSR_1C*(Percent_of_Threshold_Impedence)/100)

LV_TS_path = r'C:\LV_TS_Report'
if not os.path.exists(LV_TS_path):
    os.makedirs(LV_TS_path)

LV_TS_path_Support_File = r'C:\LV_TS_Report\Support_Files'
if not os.path.exists(LV_TS_path_Support_File):
    os.makedirs(LV_TS_path_Support_File)


LV_Imp_BattV_Dep_Age_Daily_QC = r'C:\LV_TS_Report\LV_TS_Report-Daily'
if not os.path.exists(LV_Imp_BattV_Dep_Age_Daily_QC):
    os.makedirs(LV_Imp_BattV_Dep_Age_Daily_QC)

LV_TS_path_Accumulated_QC = r'C:\LV_TS_Report\LV_Report-Accumulated'
if not os.path.exists(LV_TS_path_Accumulated_QC):
    os.makedirs(LV_TS_path_Accumulated_QC)

LV_QC_Statistics_Daily = r'C:\LV_TS_Report\LV_QC_Statistics-Daily'
if not os.path.exists(LV_QC_Statistics_Daily):
    os.makedirs(LV_QC_Statistics_Daily)

fileList = glob.glob("*.csv")
dfList = []
LVList = []

for filename in fileList:
    df                   = pd.read_csv(filename, sep=',',low_memory=False)
    df                   = df.iloc[:,:]
    CaseSN               = df.loc[:,'Case SN']
    DeviceType           = df.loc[:,' Device Type']
    Deployment_Time      = df.loc[:, 'Deployment Time (UTC)']
    Line                 = df.loc[:,'Line']
    Station              = df.loc[:,' Station']
    BattVoltage          = df.loc[:,' BatteryVoltage']
    Ch1Impedance         = df.loc[:,'Ch1 Impedance(Ohms)']
    Ch2Impedance         = df.loc[:,'Ch2 Impedance(Ohms)']
    Ch3Impedance         = df.loc[:,'Ch3 Impedance(Ohms)']
    LastScanTime         = df.loc[:,' Last Scan Time (UTC)']
    LineViewerName       = df.loc[:,'LineViewerName']
    Recording_State      = df.loc[:,'Record State']
    ScriptChecksum       = df.loc[:,'Script Checksum']
    ParamChecksum        = df.loc[:,' Param Checksum']
    TestResultsHealth    = df.loc[:,'TestResultsHealth']
    GpsHealth            = df.loc[:,'GpsHealth']
    MemoryHealth         = df.loc[:,'MemoryHealth']
    GsrHealth            = df.loc[:,'GsrHealth']
    CurrentActivity      = df.loc[:,'CurrentActivity']
    Latitude             = df.loc[:,'Latitude (DecDeg)']
    Longitude            = df.loc[:,' Longitude (DecDeg)']
    Altitude             = df.loc[:,' Altitude (m)']

    column_names = [CaseSN, DeviceType, Deployment_Time, Line, Station, BattVoltage,
                    Ch1Impedance, Ch2Impedance, Ch3Impedance, LastScanTime, LineViewerName,
                    Recording_State,ScriptChecksum,ParamChecksum,TestResultsHealth,GpsHealth,
                    MemoryHealth,GsrHealth,CurrentActivity,Latitude,Longitude,Altitude]
    catdf = pd.concat (column_names,axis=1,ignore_index =True)
    dfList.append(catdf)
    LVList.append (filename)

# Combine_All_LV_Files and Selected Column
concatDf = pd.concat(dfList,axis=0)
concatDf.rename(columns={0:'CaseSN', 1:'DeviceType', 2:'Deployment_Time', 3:'Line', 4:'Station', 5:'BattVoltage',
                         6:'Ch1Impedence',7:'Ch2Impedence',8:'Ch3Impedence',9:'LastScanTime',10:'LineViewerName',
                         11:'Recording_State',12:'ScriptChecksum',13:'ParamChecksum',14:'TestResultsHealth',
                         15:'GpsHealth',16:'MemoryHealth',17:'GsrHealth',18:'CurrentActivity',
                         19:'Latitude',20:'Longitude',21:'Altitude'},inplace = True)
LV_Rep = pd.DataFrame(concatDf)
LV_Rep["QC_Comments"]     = LV_Rep.shape[0]*[" "]

# Export_Combined Report
os.chdir = ("C:\\LV_TS_Report")
outfile_Modified =("C:\\LV_TS_Report\\Support_Files\\Combined_LV_Report.csv")
LV_Rep.to_csv(outfile_Modified,index=None)

# LV QC Total Accomplished-Modified
LV_QC_Valid  = LV_Rep.loc[:,['Line','Station','CaseSN','DeviceType','Deployment_Time','LastScanTime',
                             'BattVoltage','LineViewerName','Ch1Impedence','Ch2Impedence','Ch3Impedence',
                             'Latitude','Longitude','Altitude',
                             'Recording_State','ScriptChecksum','ParamChecksum','TestResultsHealth',
                             'GpsHealth','MemoryHealth','GsrHealth','CurrentActivity','QC_Comments']]
LV_QC_Valid  = LV_QC_Valid[pd.to_numeric(LV_QC_Valid.Line,   errors='coerce').notnull()]
LV_QC_Valid  = LV_QC_Valid[pd.to_numeric(LV_QC_Valid.Station,errors='coerce').notnull()]
LV_QC_Valid  = LV_QC_Valid[pd.to_numeric(LV_QC_Valid.CaseSN, errors='coerce').notnull()]

LV_QC_Valid['Line']    = LV_QC_Valid.Line.astype (int)
LV_QC_Valid['Station'] = LV_QC_Valid.Station.astype (int)
LV_QC_Valid['CaseSN']  = LV_QC_Valid.CaseSN.astype (int)

LV_QC_accomp_Detailed_Rep = pd.DataFrame(LV_QC_Valid)
outfile_LV_QC_accomp_Detailed_Rep   =("C:\\LV_TS_Report\\Support_Files\\LV_TS_accomplished_Report-Detailed.csv")
LV_QC_accomp_Detailed_Rep.to_csv(outfile_LV_QC_accomp_Detailed_Rep,index=None)

# LV QC Report without invalid line station and case serial number
LV_QC_accomp_valid_LN_ST = pd.DataFrame(LV_QC_Valid)
LV_QC_accomp_valid_LN_ST = LV_QC_accomp_valid_LN_ST[(LV_QC_accomp_valid_LN_ST.Line != -1)&
                          (LV_QC_accomp_valid_LN_ST.Station != -1)]

LV_QC_accomp_valid_LN_ST['Line_Station_Combined'] = (LV_QC_accomp_valid_LN_ST['Line'].map(str)+LV_QC_accomp_valid_LN_ST['Station'].map(str))
LV_QC_accomp_valid_LN_ST['Line_Station_Combined']  = LV_QC_accomp_valid_LN_ST.Line_Station_Combined.astype (float)
LV_QC_accomp_valid_LN_ST  = LV_QC_accomp_valid_LN_ST.loc[:,
                             ['Line','Station','Line_Station_Combined','CaseSN','DeviceType',
                              'Deployment_Time','LastScanTime','BattVoltage','LineViewerName',
                              'Ch1Impedence','Ch2Impedence','Ch3Impedence','Latitude','Longitude','Altitude',                              
                              'Recording_State','ScriptChecksum','ParamChecksum','TestResultsHealth',
                              'GpsHealth','MemoryHealth','GsrHealth','CurrentActivity','QC_Comments']]

outfile_LV_QC_accomp_valid_LN_ST   =("C:\\LV_TS_Report\\Support_Files\\LV_TS_accomplished_Valid_Line_Station.csv")
LV_QC_accomp_valid_LN_ST.to_csv(outfile_LV_QC_accomp_valid_LN_ST,index=None)

# LV QC Report for Script Parameters and Test Results status

LV_QC_accomp_invalid  = LV_QC_accomp_valid_LN_ST.loc[:,
                       ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time','LastScanTime',
                        'BattVoltage','Ch1Impedence','Ch2Impedence','Ch3Impedence','Latitude','Longitude','Altitude',
                        'Recording_State','ScriptChecksum','ParamChecksum','TestResultsHealth',
                        'GpsHealth','MemoryHealth','GsrHealth','CurrentActivity','QC_Comments']]

LV_QC_accomp_invalid  =  LV_QC_accomp_invalid[
                                (LV_QC_accomp_invalid.Station >= SPREAD_min_Station)&
                                (LV_QC_accomp_invalid.Station < SPREAD_max_Station)]

def trans_recroding_State(a):
    if a != ' Currently Recording':
        return 0
    elif a == ' Currently Recording':
        return 100000
    else:
        return a

def trans_CurrentActivity(b):
    if b != ' Recording':
        return 0
    elif b == ' Recording':
        return 100000
    else:
        return b

def trans_ScriptChecksum(c):
    if c != Script_Checksum:
        return 0
    elif c == Script_Checksum:
        return 100000
    else:
        return c

def trans_ParamChecksum(d):
    if d != Parameter_Checksum:
        return 0
    elif d == Parameter_Checksum:
        return 100000
    else:
        return d

def trans_Fatal(e):
    if e == " FATAL":
        return 0
    elif e != " FATAL":
        return 100000
    else:
        return e

LV_QC_accomp_invalid['Recording_State_M']   = (LV_QC_accomp_invalid['Recording_State'].apply(trans_recroding_State)).astype(int)
LV_QC_accomp_invalid['CurrentActivity_M']   = (LV_QC_accomp_invalid['CurrentActivity'].apply(trans_CurrentActivity)).astype(int)
LV_QC_accomp_invalid['ScriptChecksum_M']    = (LV_QC_accomp_invalid['ScriptChecksum'].apply(trans_ScriptChecksum)).astype(int)
LV_QC_accomp_invalid['ParamChecksum_M']     = (LV_QC_accomp_invalid['ParamChecksum'].apply(trans_ParamChecksum)).astype(int)
LV_QC_accomp_invalid['GpsHealth_M']         = (LV_QC_accomp_invalid['GpsHealth'].apply(trans_Fatal)).astype(int)
LV_QC_accomp_invalid['MemoryHealth_M']      = (LV_QC_accomp_invalid['MemoryHealth'].apply(trans_Fatal)).astype(int)
LV_QC_accomp_invalid['GsrHealth_M']         = (LV_QC_accomp_invalid['GsrHealth'].apply(trans_Fatal)).astype(int)
LV_QC_accomp_invalid['TestResultsHealth_M'] = (LV_QC_accomp_invalid['TestResultsHealth'].apply(trans_Fatal)).astype(int)

LV_QC_accomp_invalid  =  LV_QC_accomp_invalid[
                        (LV_QC_accomp_invalid.Recording_State_M < 100000)|
                        (LV_QC_accomp_invalid.CurrentActivity_M < 100000)| 
                        (LV_QC_accomp_invalid.ScriptChecksum_M  < 100000)|
                        (LV_QC_accomp_invalid.ParamChecksum_M   < 100000)|
                        (LV_QC_accomp_invalid.GpsHealth_M       < 100000)|                        
                        (LV_QC_accomp_invalid.MemoryHealth_M    < 100000)|                        
                        (LV_QC_accomp_invalid.GsrHealth_M       < 100000)|                        
                        (LV_QC_accomp_invalid.TestResultsHealth_M <100000)]

LV_QC_accomp_invalid = LV_QC_accomp_invalid.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_QC_accomp_invalid  = LV_QC_accomp_invalid.loc[:,
                        ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time','LastScanTime',
                        'Latitude','Longitude','Altitude','Recording_State','ScriptChecksum','ParamChecksum','TestResultsHealth',
                        'GpsHealth','MemoryHealth','GsrHealth','CurrentActivity','QC_Comments']]

outfile_Modified =("C:\\LV_TS_Report\\Support_Files\\LV_Total_Spread_QC_Fail_Status.csv")
LV_QC_accomp_invalid.to_csv(outfile_Modified,index=None)	

# LV QC Total Accomplished_Report-with 'Line','Station','CaseSN'
LV_QC_Valid_P  =  LV_QC_accomp_valid_LN_ST.loc[:,['Line','Station','Line_Station_Combined','CaseSN']]
LV_QC_Valid_P  =  LV_QC_Valid_P[(LV_QC_Valid_P.Line != -1)&
                                (LV_QC_Valid_P.Station >= SPREAD_min_Station)&
                                (LV_QC_Valid_P.Station < SPREAD_max_Station)&
                                (LV_QC_Valid_P.CaseSN != -1)]
LV_QC_Valid_P    = LV_QC_Valid_P.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_QC_accomplished_Rep = pd.DataFrame(LV_QC_Valid_P)

outfile_Modified =("C:\\LV_TS_Report\\Support_Files\\LV_TS_accomplished_Report.csv")
LV_QC_accomplished_Rep.to_csv(outfile_Modified,index=None)	

# LV QC Statistics
Line_Station_Minimum    = LV_QC_accomplished_Rep.groupby('Line').Station.min()
Line_Station_Maximum    = LV_QC_accomplished_Rep.groupby('Line').Station.max()
Station_Count           = LV_QC_accomplished_Rep.groupby('Line').Station.count()
LV_QC_Done              = [Line_Station_Minimum,Line_Station_Maximum,Station_Count]
LV_QC_Done_Combined     = pd.concat (LV_QC_Done,axis=1,ignore_index =True)
LV_QC_Done_Combined.reset_index(inplace=True)
LV_QC_Done_Combined.rename(columns={0:'LV_Start_Station', 1:'LV_End_Station',2:'LV_Station_Count'},inplace = True)

outfile_LV_QC_Stat =("C:\\LV_TS_Report\\Support_Files\\LV_QC_Statistics.csv")
LV_QC_Done_Combined.to_csv(outfile_LV_QC_Stat,index=None)

# Generate LV QC Total planned RPS
QCList = []
List_RLine    = list(LV_QC_Done_Combined.Line)

for i in range(len(List_RLine)):
        ListRL  = List_RLine[i]
        List_St = list(range((LV_QC_Done_Combined.LV_Start_Station[i]),(LV_QC_Done_Combined.LV_End_Station[i]+1),Rec_Station_increment))
        QC_List = {'RPSLine': ListRL, 'RPSStation': List_St}
        QC_DF   = pd.DataFrame(data=QC_List,index=None)
        QC_DF1  = QC_DF.iloc[:,:]
        RPSLine    = QC_DF1.loc[:,'RPSLine']
        RPSStation = QC_DF1.loc[:,'RPSStation']
        LN_ST   = [RPSLine,RPSStation]
        QCcatdf = pd.concat (LN_ST,axis=1,ignore_index =True)
        QCList.append(QCcatdf)

concatQCList = pd.concat(QCList,axis=0)
concatQCList.rename(columns={0:'Line', 1:'Station'},inplace = True)
LV_QC_Planned_Rep = pd.DataFrame(concatQCList)
LV_QC_Planned_Rep['Line_Station_Combined'] = (LV_QC_Planned_Rep['Line'].map(str)+LV_QC_Planned_Rep['Station'].map(str)).astype(int)
outfile_Modified =("C:\\LV_TS_Report\\Support_Files\\LV_TS_Planned_Report.csv")
LV_QC_Planned_Rep.to_csv(outfile_Modified,index=None)

# Generate LV QC Missing Report
LV_QC_Missing_Rep = pd.merge(LV_QC_Planned_Rep, LV_QC_accomplished_Rep,
                    how='left', on='Line_Station_Combined',
                    suffixes=('_LV_Planned', '_LV_Accomplished'))
LV_QC_Missing_Rep_Detail = LV_QC_Missing_Rep.fillna('Missing_LV_QC')

outfile_Modified =("C:\\LV_TS_Report\\Support_Files\\LV_QC_Detail_Report.csv")
LV_QC_Missing_Rep_Detail.to_csv(outfile_Modified,index=None)

LV_QC_Missing_Rep = LV_QC_Missing_Rep[(LV_QC_Missing_Rep.Line_LV_Accomplished.isnull())|
                                      (LV_QC_Missing_Rep.Station_LV_Accomplished.isnull())]

LV_QC_Missing_Rep            = LV_QC_Missing_Rep.loc[:,['Line_LV_Planned','Station_LV_Planned','Line_Station_Combined']]
LV_QC_Missing_Rep["LV_Flag"] = LV_QC_Missing_Rep.shape[0]*["Missing Line Viewing QC "]


outfile_LV_QC_Missing_Report = ("C:\\LV_TS_Report\\Support_Files\\LV_QC_Missing_Report.csv")
LV_QC_Missing_Rep.to_csv(outfile_LV_QC_Missing_Report,index=None)

# Generate LV QC Unknown Station Report
LV_QC_Null_P     = (LV_QC_accomp_Detailed_Rep.loc[:,['Line','Station','CaseSN']]).astype(int)
LV_QC_Null_P     = LV_QC_Null_P[(LV_QC_Null_P.Line == -1)|
                                (LV_QC_Null_P.Station < SPREAD_min_Station)|
                                (LV_QC_Null_P.Station > SPREAD_max_Station)]
LV_Null_Report   = pd.merge(LV_QC_Null_P, LV_QC_accomplished_Rep, how='left', on='CaseSN')
LV_Null_Report   = LV_Null_Report[(LV_Null_Report.Line_y.isnull())|
                                 (LV_Null_Report.Station_y.isnull())]
LV_Null_Report   = LV_Null_Report.loc[:,['Line_x','Station_x','CaseSN']]

LV_Null_Report["LV_Flag"] = LV_Null_Report.shape[0]*["Unknowun _CaseSN_Line_Station"]
outfile_LV_Null_Report    =("C:\\LV_TS_Report\\Support_Files\\LV_QC_Null_Report.csv")
LV_Null_Report.to_csv(outfile_LV_Null_Report,index=None)

## Exporting Daily Excel LV_QC_Statistics-Daily ####

from datetime import datetime

def get_LV_QC_Statistics_Report_datetime():
    return "LV_QC_Statistics for Line_Coordinator-" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"

LV_QC_Statistics_TS_Report = get_LV_QC_Statistics_Report_datetime()
Daily_LV_QC_Statistics_path      = "C:\\LV_TS_Report\\LV_QC_Statistics-Daily\\"+ LV_QC_Statistics_TS_Report
XLSX_writer_QC_Stat = pd.ExcelWriter(Daily_LV_QC_Statistics_path)

LV_QC_Done_Combined.to_excel(XLSX_writer_QC_Stat,'LV_QC_Stat',index=False)
LV_QC_Missing_Rep.to_excel(XLSX_writer_QC_Stat,'LV_Missing_QC',index=False)
LV_QC_Missing_Rep_Detail.to_excel(XLSX_writer_QC_Stat,'LV_QC_Details',index=False)
LV_Null_Report.to_excel(XLSX_writer_QC_Stat,'LV_Unknown',index=False)

XLSX_writer_QC_Stat.save()
XLSX_writer_QC_Stat.close()

########## Deployment Age Calculation###############
LV_Rep_Deploy_Time  = pd.DataFrame(LV_QC_accomp_valid_LN_ST)
LV_Rep_Deploy_Time  =  LV_Rep_Deploy_Time[
                                (LV_Rep_Deploy_Time.Station >= SPREAD_min_Station)&
                                (LV_Rep_Deploy_Time.Station < SPREAD_max_Station)
                                ]
LV_Rep_Deploy_Time = LV_Rep_Deploy_Time[(LV_Rep_Deploy_Time.DeviceType == ' 1-C GSR')|
                                        (LV_Rep_Deploy_Time.DeviceType == ' 1-C GSX')|
                                        (LV_Rep_Deploy_Time.DeviceType == ' GSR X3')]

LV_Rep_Deploy_Time = LV_Rep_Deploy_Time[(LV_Rep_Deploy_Time.Deployment_Time != 'Unknown')&
                                        (LV_Rep_Deploy_Time.Deployment_Time != ' Unknown')&
                                        (LV_Rep_Deploy_Time.Deployment_Time != ' -')&
                                        (LV_Rep_Deploy_Time.Deployment_Time != '-')&
                                        (LV_Rep_Deploy_Time.Deployment_Time != ' ?')&
                                        (LV_Rep_Deploy_Time.Deployment_Time != '?')]

LV_Rep_Deploy_Time['Deployment_Time']= pd.to_datetime (LV_Rep_Deploy_Time.Deployment_Time)
LV_Rep_Deploy_Time['DeploymentAge']  =(Default_Date - LV_Rep_Deploy_Time['Deployment_Time'])
LV_Rep_Deploy_Time['DeploymentAge_DD HH-MM-SS']  =(Default_Date - LV_Rep_Deploy_Time['Deployment_Time']).astype(str)
LV_Rep_Deploy_Time['DeploymentAge_DD HH-MM-SS'] = LV_Rep_Deploy_Time['DeploymentAge_DD HH-MM-SS'].str.slice(0,15)

LV_Rep_Deploy_Time['DeploymentAge']  = LV_Rep_Deploy_Time['DeploymentAge'].dt.days
LV_Rep_Deploy_Time['DeploymentAge']  = LV_Rep_Deploy_Time.DeploymentAge.astype (float)
LV_Rep_Deploy_Time = LV_Rep_Deploy_Time[LV_Rep_Deploy_Time.DeploymentAge >= Deployment_Age]
LV_Rep_Deploy_Time = LV_Rep_Deploy_Time.loc[:,
                    ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time','LastScanTime', 
                     'BattVoltage','Ch1Impedence', 'Ch2Impedence','Ch3Impedence','Latitude','Longitude','Altitude', 
                     'LineViewerName','DeploymentAge_DD HH-MM-SS','QC_Comments']]

LV_Rep_Deploy_Time = LV_Rep_Deploy_Time.drop_duplicates(['Line_Station_Combined'],keep='last')

# Deployment Age Export Report
outfile_Deployment_Age_FAIL =("C:\\LV_TS_Report\\Support_Files\\Deployment Age_FAIL_LV_Report.csv")
LV_Rep_Deploy_Time.to_csv(outfile_Deployment_Age_FAIL,index=None)


########## End of Deployment Age Calculation###############

# Battery Voltage Fail Check############## 
LV_Rep_BattVoltage_Check  = pd.DataFrame(LV_QC_accomp_valid_LN_ST)
LV_Rep_BattVoltage_Check  =  LV_Rep_BattVoltage_Check[
                                (LV_Rep_BattVoltage_Check.Station >= SPREAD_min_Station)&
                                (LV_Rep_BattVoltage_Check.Station < SPREAD_max_Station)]
LV_Rep_BattVoltage_Check = LV_Rep_BattVoltage_Check[(LV_Rep_BattVoltage_Check.DeviceType == ' 1-C GSR')|
                                        (LV_Rep_BattVoltage_Check.DeviceType == ' 1-C GSX')|
                                        (LV_Rep_BattVoltage_Check.DeviceType == ' GSR X3')]

def trans_Batt_Volt_Change(y):
    if y == ' OPEN':
        return 0
    elif y == 'OPEN':
        return 0
    elif y == 'Off':
        return 0
    elif y == ' Off':
        return 0
    elif y == ' ?':
        return 0
    elif y == '?':
        return 0
    elif y == '-':
        return 0
    elif y == ' -':
        return 0
    elif y == 'Unknown':
        return 0
    elif y == ' Unknown':
        return 0
    else:
        return y

LV_Rep_BattVoltage_Check['Batt_Voltage'] = LV_Rep_BattVoltage_Check['BattVoltage'].apply(trans_Batt_Volt_Change)
LV_Rep_BattVoltage_Check['Batt_Voltage'] = LV_Rep_BattVoltage_Check.Batt_Voltage.astype (float)

LV_Rep_BattVoltage_Fail = pd.DataFrame(LV_Rep_BattVoltage_Check)
LV_Rep_BattVoltage_Pass = pd.DataFrame(LV_Rep_BattVoltage_Check)

LV_Rep_BattVoltage_Fail = LV_Rep_BattVoltage_Fail[LV_Rep_BattVoltage_Fail.Batt_Voltage <= DesiredBattVoltage]
LV_Rep_BattVoltage_Pass = LV_Rep_BattVoltage_Pass[LV_Rep_BattVoltage_Pass.Batt_Voltage > DesiredBattVoltage]
LV_Rep_BattVoltage_Fail = LV_Rep_BattVoltage_Fail.loc[:,
                    ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','BattVoltage','Batt_Voltage','Deployment_Time','LastScanTime',
                     'Ch1Impedence', 'Ch2Impedence','Ch3Impedence','Latitude','Longitude','Altitude','LineViewerName' ,'QC_Comments']]

LV_Rep_BattVoltage_Pass = LV_Rep_BattVoltage_Pass.loc[:,
                    ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','BattVoltage','Batt_Voltage','Deployment_Time','LastScanTime',
                     'Ch1Impedence', 'Ch2Impedence','Ch3Impedence','Latitude','Longitude','Altitude','LineViewerName','QC_Comments']]
LV_Rep_BattVoltage_Fail = LV_Rep_BattVoltage_Fail.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_Rep_BattVoltage_Pass = LV_Rep_BattVoltage_Pass.drop_duplicates(['Line_Station_Combined'],keep='last')

LV_Rep_BattVoltage_Fail_M = pd.merge(LV_Rep_BattVoltage_Fail, LV_Rep_BattVoltage_Pass, how='left', on=['Line_Station_Combined','Line_Station_Combined'])
LV_Rep_BattVoltage_Fail_M.drop(columns=['Batt_Voltage_y','Line_y','Station_y','DeviceType_y','Deployment_Time_y',
                                        'Latitude_y','Longitude_y','Altitude_y', 'QC_Comments_y'],axis=1,inplace=True)

LV_Rep_BattVoltage_Fail_M["QC_Comments_x"] = LV_Rep_BattVoltage_Fail_M.shape[0]*["Right- After_QC"]

### Client Statistics Report- Batt_Volt

# Check Batt Volt Pass_Stats
Batt_Volt_RNG_Six = LV_Rep_BattVoltage_Pass[(LV_Rep_BattVoltage_Pass.Batt_Voltage > 16.5)].count()
Batt_Volt_RNG_Six = Batt_Volt_RNG_Six['Line_Station_Combined']

Batt_Volt_RNG_Five = LV_Rep_BattVoltage_Pass[(LV_Rep_BattVoltage_Pass.Batt_Voltage > 16.4)&
                                (LV_Rep_BattVoltage_Pass.Batt_Voltage <= 16.5)].count()
Batt_Volt_RNG_Five = Batt_Volt_RNG_Five['Line_Station_Combined']

Batt_Volt_RNG_Four = LV_Rep_BattVoltage_Pass[(LV_Rep_BattVoltage_Pass.Batt_Voltage > 16.3)&
                                (LV_Rep_BattVoltage_Pass.Batt_Voltage <= 16.4)].count()
Batt_Volt_RNG_Four = Batt_Volt_RNG_Four['Line_Station_Combined']

Batt_Volt_RNG_Three = LV_Rep_BattVoltage_Pass[(LV_Rep_BattVoltage_Pass.Batt_Voltage > 16.2)&
                                (LV_Rep_BattVoltage_Pass.Batt_Voltage <= 16.3)].count()
Batt_Volt_RNG_Three = Batt_Volt_RNG_Three['Line_Station_Combined']

Batt_Volt_RNG_Two   = LV_Rep_BattVoltage_Pass[(LV_Rep_BattVoltage_Pass.Batt_Voltage > 16.0)&
                                (LV_Rep_BattVoltage_Pass.Batt_Voltage <= 16.2)].count()
Batt_Volt_RNG_Two   = Batt_Volt_RNG_Two['Line_Station_Combined']

# Check Batt Volt Fail_Stats
LV_Rep_BattVoltage_Fail_Client   = LV_Rep_BattVoltage_Fail_M[(LV_Rep_BattVoltage_Fail_M.CaseSN_y.isnull())]

Batt_Volt_RNG_One = LV_Rep_BattVoltage_Fail_Client[(LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x > 15.9)&
                                (LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x <= 16.0)].count()
Batt_Volt_RNG_One = Batt_Volt_RNG_One['Line_Station_Combined']
Batt_Volt_RNG_Zero = LV_Rep_BattVoltage_Fail_Client[(LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x > 15.7)&
                                                    (LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x <= 15.9)].count()
Batt_Volt_RNG_Zero = Batt_Volt_RNG_Zero['Line_Station_Combined']

Batt_Volt_RNG_MOne = LV_Rep_BattVoltage_Fail_Client[(LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x > 15.5)&
                                                    (LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x <= 15.7)].count()
Batt_Volt_RNG_MOne = Batt_Volt_RNG_MOne['Line_Station_Combined']

Batt_Volt_RNG_Mtwo = LV_Rep_BattVoltage_Fail_Client[(LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x > 15.3)&
                                                    (LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x <= 15.5)].count()
Batt_Volt_RNG_Mtwo = Batt_Volt_RNG_Mtwo['Line_Station_Combined']

Batt_Volt_RNG_Mthree = LV_Rep_BattVoltage_Fail_Client[(LV_Rep_BattVoltage_Fail_Client.Batt_Voltage_x <=  15.3)].count()
                                                    
Batt_Volt_RNG_Mthree = Batt_Volt_RNG_Mthree['Line_Station_Combined']

LV_QC_Batt_Stat_Client = pd.DataFrame({'Batt_Volt: <= 15.30':[Batt_Volt_RNG_Mthree],
                                       'Batt_Volt: >15.30 & <= 15.50':[Batt_Volt_RNG_Mtwo],
                                       'Batt_Volt: >15.50 & <= 15.70':[Batt_Volt_RNG_MOne],
                                       'Batt_Volt: >15.70 & <= 15.90':[Batt_Volt_RNG_Zero],
                                       'Batt_Volt: >15.90 & <= 16.0':[Batt_Volt_RNG_One],
                                       'Batt_Volt: >16.00 & <= 16.20':[Batt_Volt_RNG_Two],
                                       'Batt_Volt: >16.20 & <= 16.30':[Batt_Volt_RNG_Three],                                   
                                       'Batt_Volt: >16.30 & <= 16.40':[Batt_Volt_RNG_Four],
                                       'Batt_Volt: >16.40 & <= 16.50':[Batt_Volt_RNG_Five],
                                       'Batt_Volt: > 16.5':[Batt_Volt_RNG_Six]},index=None)

LV_QC_Batt_Stat_Client= LV_QC_Batt_Stat_Client.T
LV_QC_Batt_Stat_Client.rename(columns = {0:'Total_Count'},inplace = True)

### Battery Voltage Fail/Pass Export Report
LV_Rep_BattVoltage_Fail_M.drop(columns=['Batt_Voltage_x'],axis=1,inplace=True)
outfile_BattVoltage_FAIL =("C:\\LV_TS_Report\\Support_Files\\BattVoltage_FAIL_LV_Report.csv")
LV_Rep_BattVoltage_Fail_M.to_csv(outfile_BattVoltage_FAIL,index=None)

outfile_BattVoltage_PASS =("C:\\LV_TS_Report\\Support_Files\\BattVoltage_PASS_LV_Report.csv")
LV_Rep_BattVoltage_Pass.to_csv(outfile_BattVoltage_PASS,index=None)

# Impedence Fail Check ############## Impedence Fail Check ###########
LV_Rep  = pd.DataFrame(LV_QC_accomp_valid_LN_ST)
LV_Rep  =  LV_Rep[(LV_Rep.Station >= SPREAD_min_Station)& (LV_Rep.Station < SPREAD_max_Station)]
                                
def trans_Ch_Impedance_Change(x):
    if x == ' OPEN':
        return 99999999
    elif x == 'OPEN':
        return 99999999
    elif x == 'Off':
        return 88888888
    elif x == ' Off':
        return 88888888
    elif x == ' ?':
        return 77777777
    elif x == '?':
        return 77777777
    elif x == '-':
        return 66666666
    elif x == ' -':
        return 66666666
    elif x == 'Unknown':
        return 66666666
    elif x == ' Unknown':
        return 66666666
    else:
        return x      

# Apply filter to change OPEN, off, -, NA values..... xxxxxxx
LV_Rep['Ch1_Impedance'] = LV_Rep['Ch1Impedence'].apply(trans_Ch_Impedance_Change)
LV_Rep['Ch2_Impedance'] = LV_Rep['Ch2Impedence'].apply(trans_Ch_Impedance_Change)
LV_Rep['Ch3_Impedance'] = LV_Rep['Ch3Impedence'].apply(trans_Ch_Impedance_Change)

outfile_MOD_Station =("C:\\LV_TS_Report\\Support_Files\\MOD_Station_LV_Report.csv")
LV_Rep.to_csv(outfile_MOD_Station,index=None)

# Channge 'Ch1_Impedance','Ch2_Impedance','Ch3_Impedance' to object to int
LV_Rep['Ch1_Impedance'] = LV_Rep.Ch1_Impedance.astype (int)
LV_Rep['Ch2_Impedance'] = LV_Rep.Ch2_Impedance.astype (int)
LV_Rep['Ch3_Impedance'] = LV_Rep.Ch3_Impedance.astype (int)

# Search 'Ch1_Impedance','Ch2_Impedance','Ch3_Impedance' impedence FAILED
LV_Impedence_FAIL_GSRX3  =    LV_Rep[(LV_Rep.Ch1_Impedance <Low_Threshold_Imp_GSRX3)|(LV_Rep.Ch1_Impedance >High_Threshold_Imp_GSRX3)|
                                     (LV_Rep.Ch2_Impedance <Low_Threshold_Imp_GSRX3)|(LV_Rep.Ch2_Impedance >High_Threshold_Imp_GSRX3)|
                                     (LV_Rep.Ch3_Impedance <Low_Threshold_Imp_GSRX3)|(LV_Rep.Ch3_Impedance >High_Threshold_Imp_GSRX3)]
LV_Impedence_FAIL_GSRX1  =    LV_Rep[(LV_Rep.Ch1_Impedance <Low_Threshold_Imp_GSR_1C)|(LV_Rep.Ch1_Impedance >High_Threshold_Imp_GSR_1C)]
                                                                    

LV_Impedence_PASS_GSRX3  =    LV_Rep[(LV_Rep.Ch1_Impedance >Low_Threshold_Imp_GSRX3)&(LV_Rep.Ch1_Impedance <High_Threshold_Imp_GSRX3)&
                                     (LV_Rep.Ch2_Impedance >Low_Threshold_Imp_GSRX3)&(LV_Rep.Ch2_Impedance <High_Threshold_Imp_GSRX3)&
                                     (LV_Rep.Ch3_Impedance >Low_Threshold_Imp_GSRX3)&(LV_Rep.Ch3_Impedance <High_Threshold_Imp_GSRX3)]

LV_Impedence_PASS_GSRX1  =    LV_Rep[(LV_Rep.Ch1_Impedance >Low_Threshold_Imp_GSR_1C)&(LV_Rep.Ch1_Impedance <High_Threshold_Imp_GSR_1C)]

# Filtering GSR-X3 and GSR-X1
LV_Impedence_FAIL_GSR_X3 = LV_Impedence_FAIL_GSRX3.loc[:,
                           ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time', 'LastScanTime',
                            'BattVoltage','Ch1Impedence', 'Ch2Impedence','Ch3Impedence', 'LineViewerName',
                            'Latitude','Longitude','Altitude','QC_Comments','Ch1_Impedance','Ch2_Impedance','Ch3_Impedance']]
LV_Impedence_PASS_GSRX3 = LV_Impedence_PASS_GSRX3.loc[:,
                           ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time', 'LastScanTime',
                            'BattVoltage','Ch1Impedence', 'Ch2Impedence','Ch3Impedence', 'LineViewerName',
                            'Latitude','Longitude','Altitude','QC_Comments']]
LV_Impedence_FAIL_GSR_X1 = LV_Impedence_FAIL_GSRX1.loc[:,
                           ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time','LastScanTime',
                            'BattVoltage','Ch1Impedence','LineViewerName','Latitude','Longitude','Altitude',
                            'QC_Comments','Ch1_Impedance']]

LV_Impedence_PASS_GSRX1 = LV_Impedence_PASS_GSRX1.loc[:,
                           ['Line','Station','Line_Station_Combined','CaseSN','DeviceType','Deployment_Time','LastScanTime',
                            'BattVoltage','Ch1Impedence','LineViewerName','Latitude','Longitude','Altitude','QC_Comments']]
                            
LV_Impedence_FAIL_GSR_X3   = LV_Impedence_FAIL_GSR_X3[LV_Impedence_FAIL_GSR_X3.DeviceType == ' GSR X3']
LV_Impedence_PASS_GSRX3    = LV_Impedence_PASS_GSRX3 [LV_Impedence_PASS_GSRX3.DeviceType == ' GSR X3']

LV_Impedence_FAIL_GSR_X3   = LV_Impedence_FAIL_GSR_X3.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_Impedence_PASS_GSRX3    = LV_Impedence_PASS_GSRX3.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_Impedence_FAIL_GSR_X3_M = pd.merge(LV_Impedence_FAIL_GSR_X3, LV_Impedence_PASS_GSRX3, how='left', on=['Line_Station_Combined','Line_Station_Combined'])

LV_Impedence_FAIL_GSR_X3_M.drop(columns = ['Line_y','Station_y','DeviceType_y','Deployment_Time_y',
                                           'Latitude_y','Longitude_y','Altitude_y','QC_Comments_y'],axis=1,inplace=True)
LV_Impedence_FAIL_GSR_X3_M["QC_Comments_x"] = LV_Impedence_FAIL_GSR_X3_M.shape[0]*["Right- After_QC"]

LV_Impedence_FAIL_GSR_X1 = LV_Impedence_FAIL_GSR_X1[(LV_Impedence_FAIL_GSR_X1.DeviceType == ' 1-C GSR')|(LV_Impedence_FAIL_GSR_X1.DeviceType == ' 1-C GSX')]
LV_Impedence_FAIL_GSR_X1 = LV_Impedence_FAIL_GSR_X1.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_Impedence_PASS_GSRX1  = LV_Impedence_PASS_GSRX1.drop_duplicates(['Line_Station_Combined'],keep='last')
LV_Impedence_FAIL_GSR_X1_M = pd.merge(LV_Impedence_FAIL_GSR_X1, LV_Impedence_PASS_GSRX1, how='left', on=['Line_Station_Combined','Line_Station_Combined'])
LV_Impedence_FAIL_GSR_X1_M.drop(columns =  ['Line_y','Station_y','DeviceType_y','Deployment_Time_y',
                                            'Latitude_y','Longitude_y','Altitude_y', 'QC_Comments_y'],axis=1,inplace=True)
LV_Impedence_FAIL_GSR_X1_M["QC_Comments_x"] = LV_Impedence_FAIL_GSR_X1_M.shape[0]*["Right- After_QC"]

### Client Statistics Report- Impedence
GSR_X3_Impedence_RNG_Two = LV_Impedence_PASS_GSRX3.CaseSN.count()
GSR_X3_Impedence_Fail_Client   = LV_Impedence_FAIL_GSR_X3_M[(LV_Impedence_FAIL_GSR_X3_M.CaseSN_y.isnull())]

GSR_X3_Impedence_RNG_Zero = GSR_X3_Impedence_Fail_Client[(GSR_X3_Impedence_Fail_Client.Ch1_Impedance >= 66666666)|
                                (GSR_X3_Impedence_Fail_Client.Ch2_Impedance >= 66666666)|
                                (GSR_X3_Impedence_Fail_Client.Ch3_Impedance >= 66666666)].count()
GSR_X3_Impedence_RNG_Zero = GSR_X3_Impedence_RNG_Zero['Line_Station_Combined']

GSR_X3_Impedence_RNG_One = GSR_X3_Impedence_Fail_Client[(GSR_X3_Impedence_Fail_Client.Ch1_Impedance < 7128)|
                                (GSR_X3_Impedence_Fail_Client.Ch2_Impedance < 7128)|
                                (GSR_X3_Impedence_Fail_Client.Ch3_Impedance < 7128)].count()
GSR_X3_Impedence_RNG_One = GSR_X3_Impedence_RNG_One['Line_Station_Combined']

GSR_X3_Impedence_RNG_Three = GSR_X3_Impedence_Fail_Client[((GSR_X3_Impedence_Fail_Client.Ch1_Impedance > 9465)&(GSR_X3_Impedence_Fail_Client.Ch1_Impedance < 66666666))|
                                ((GSR_X3_Impedence_Fail_Client.Ch2_Impedance > 9465)&(GSR_X3_Impedence_Fail_Client.Ch2_Impedance < 66666666))|
                                ((GSR_X3_Impedence_Fail_Client.Ch3_Impedance > 9465)&(GSR_X3_Impedence_Fail_Client.Ch3_Impedance < 66666666))].count()
GSR_X3_Impedence_RNG_Three = GSR_X3_Impedence_RNG_Three['Line_Station_Combined']

GSR_X3_LV_QC_Impedence_Stat_Client = pd.DataFrame({'Impedence: = OPEN':[GSR_X3_Impedence_RNG_Zero],'Impedence: >0 & < 7128':[GSR_X3_Impedence_RNG_One],
                                   'Impedence: >= 7128 & <= 9465':[GSR_X3_Impedence_RNG_Two],'Impedence: >9465':[GSR_X3_Impedence_RNG_Three]},index=None)                                  
                                   
GSR_X3_LV_QC_Impedence_Stat_Client= GSR_X3_LV_QC_Impedence_Stat_Client.T
GSR_X3_LV_QC_Impedence_Stat_Client.rename(columns = {0:'Total_Count'},inplace = True)

# Export Impedence FAIL CSV file for All REPORTS
LV_Impedence_FAIL_GSR_X3_M.drop(columns=['Ch1_Impedance','Ch2_Impedance','Ch3_Impedance'],axis=1,inplace=True)
outfile_Impedence_FAIL_GSR_X3 =("C:\\LV_TS_Report\\Support_Files\\GSR_X3_Impedence_FAIL_LV_Report.csv")
LV_Impedence_FAIL_GSR_X3_M.to_csv(outfile_Impedence_FAIL_GSR_X3,index=None)
outfile_Impedence_PASS_GSR_X3 =("C:\\LV_TS_Report\\Support_Files\\GSR_X3_Impedence_PASS_LV_Report.csv")
LV_Impedence_PASS_GSRX3.to_csv(outfile_Impedence_PASS_GSR_X3,index=None)

LV_Impedence_FAIL_GSR_X1_M.drop(columns=['Ch1_Impedance'],axis=1,inplace=True)
outfile_Impedence_FAIL_GSR_X1 =("C:\\LV_TS_Report\\Support_Files\\GSR_X1_Impedence_FAIL_LV_Report.csv")
LV_Impedence_FAIL_GSR_X1_M.to_csv(outfile_Impedence_FAIL_GSR_X1,index=None)
outfile_Impedence_PASS_GSR_X1 =("C:\\LV_TS_Report\\Support_Files\\GSR_X1_Impedence_PASS_LV_Report.csv")
LV_Impedence_PASS_GSRX1.to_csv(outfile_Impedence_PASS_GSR_X1,index=None)

## Exporting Daily Excel LV_Daily_Imp_BattV_Dep_Age_Report ####

from datetime import datetime

def get_LV_Daily_Imp_BattV_Dep_Age_Report_datetime():
    return "LV_Daily_TS_Report-" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"

LV_Daily_Imp_BattV_Dep_Age_Report = get_LV_Daily_Imp_BattV_Dep_Age_Report_datetime()
Daily_LV_Imp_BattV_Dep_Age_path   = "C:\\LV_TS_Report\\\LV_TS_Report-Daily\\" + LV_Daily_Imp_BattV_Dep_Age_Report
XLSX_writer = pd.ExcelWriter(Daily_LV_Imp_BattV_Dep_Age_path)
LV_Impedence_FAIL_GSR_X3_M.to_excel(XLSX_writer,'Imp_FAIL_GSRX3',index=False)
LV_Impedence_FAIL_GSR_X1_M.to_excel(XLSX_writer,'Imp_FAIL_GSRX1',index=False)
LV_Rep_BattVoltage_Fail_M.to_excel(XLSX_writer,'BattVolt_FAIL',index=False)
LV_Rep_Deploy_Time.to_excel(XLSX_writer,'Deploy_Age_FAIL',index=False)
LV_QC_accomp_invalid.to_excel(XLSX_writer,'Spread_QC_Failure',index=False)
XLSX_writer.save()
XLSX_writer.close()
    
## Exporting Accumulated Excel LV TS Report ####
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# Export Accumulated_LV_Report.xlsx_COMPLETE_REPORT

if os.path.isfile("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx"):
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_accomp_invalid,sheet_name='LV_QC_Fail_Overview',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Done_Combined,sheet_name='LV_QC_Stat',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Missing_Rep,sheet_name='LV_Missing_QC',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Missing_Rep_Detail,sheet_name='LV_QC_Details',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Null_Report,sheet_name='LV_Null',header = False,index=False)    
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Impedence_FAIL_GSR_X3_M,sheet_name='Imp_FAIL_GSRX3',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Impedence_FAIL_GSR_X1_M,sheet_name='Imp_FAIL_GSRX1',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Rep_BattVoltage_Fail_M,sheet_name='BattVolt_FAIL',header = False,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Rep_Deploy_Time,sheet_name='Deploy_Age_FAIL',header = False,index=False)
else:
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_accomp_invalid,sheet_name='LV_QC_Fail_Overview',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Done_Combined,sheet_name='LV_QC_Stat',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Missing_Rep,sheet_name='LV_Missing_QC',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_QC_Missing_Rep_Detail,sheet_name='LV_QC_Details',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Null_Report,sheet_name='LV_Null',header = True,index=False)    
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Impedence_FAIL_GSR_X3_M,sheet_name='Imp_FAIL_GSRX3',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Impedence_FAIL_GSR_X1_M,sheet_name='Imp_FAIL_GSRX1',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Rep_BattVoltage_Fail_M,sheet_name='BattVolt_FAIL',header = True,index=False)
    append_df_to_excel("C:\\LV_TS_Report\\LV_Report-Accumulated\\LV_Accumulated_TS_Report.xlsx",LV_Rep_Deploy_Time,sheet_name='Deploy_Age_FAIL',header = True,index=False)

# Export Client Statistics_REPORT
from datetime import datetime
def get_LV_Daily_LV_QC_Stat_Client_datetime():
    return "LV_QC_Statistics for Client- " + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"

LV_QC_Client             = get_LV_Daily_LV_QC_Stat_Client_datetime()
LV_QC_Client_path        = "C:\\LV_TS_Report\\\LV_QC_Statistics-Daily\\" + LV_QC_Client
XLSX_writer_LV_QC_Client = pd.ExcelWriter(LV_QC_Client_path)
LV_QC_Batt_Stat_Client.to_excel(XLSX_writer_LV_QC_Client,'Batt_Stat',index=True)
GSR_X3_LV_QC_Impedence_Stat_Client.to_excel(XLSX_writer_LV_QC_Client,'GSR_X3_Imp_Stat',index=True)
XLSX_writer_LV_QC_Client.save()
XLSX_writer_LV_QC_Client.close()

